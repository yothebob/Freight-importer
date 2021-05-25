from tkinter import *
from tkinter.filedialog import askopenfilename
import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import math

file = ''

cwd = os.getcwd()
print('current dir: ',cwd)
os.chdir(cwd)


def set_file(_file):
    global file
    file= _file

def file_name():
    _path=askopenfilename(filetypes=[('Excel Files','.xlsm')])
    set_file(_path)

# load in excel estimate
file = file + '.xlsm'

def import_data():
    workbook = load_workbook(filename=file,data_only=True)

    area_a = workbook['Area A']
    area_b = workbook['Area B']
    area_c = workbook['Area C']
    area_d = workbook['Area D']

    areas = [area_a,area_b,area_c,area_d]
    items = {}
    nones = ['None','none','Item',0,None]

    for area in areas:
        for num in range(5,251):
            if area['B'+str(num)].value not in nones:
                if area['E'+str(num)].value not in nones:
                    if area['B'+str(num)].value not in items.keys():
                        items[area['B'+str(num)].value] = area['E'+str(num)].value
                    else:
                        items[area['B'+str(num)].value] += area['E'+str(num)].value
        
    print(items)


    #fill out freight form          
    freight = load_workbook(filename='Freight Calculator Model-Rework.xlsm',data_only=True,keep_vba=True)

    freight_wb = freight['Freight Quote Workbook']

    for num in range(3,109):
        if freight_wb['B'+str(num)].value in items.keys():
            freight_wb['D'+str(num)] = items[freight_wb['B'+str(num)].value]

    freight_name=input('what do you want to call your freight quote?')
    freight.save(filename=(freight_name+'.xlsm'))


    #bundles, pallets, boxes

    longparts = ['Splice MF','FI- Blk','Pocket Infill- Blk','PTE1 - Blk','SPE1- Blk','SPE2- Blk','PE ',
                 'PE-135,PEU- Blk 18 3/4','PEU500, Blk','PEU650- Blk','BR100- Blk','BR200- Blk','BR500-Blk'',TR100',
                 'TR175A  Blk','TR200 Blk','TR375 Blk','TR400 Blk','TR670 -Blk','TR999 Blk']
    halflength = ['GVS Bot Glass Vnl','GVS Top Glass Vnl','PVI','RCBE1','RCBE2']
    pallet_parts = ['CTG - CL 1/2','CTG - CL 1/4','PT1-360- Blk','PT1-420- Blk','FP45- Blk','P361- Blk U','P3613- Blk','P362- Blk U','P421- Blk U','P4213- Blk','FP56- Blk']
    box_parts = ['RCB1-12"','RCB2-12"','INT-90°/650 rail','Splice Series 200','FMPBS3- Blk','FMPBS 1- Blk','FMPBS 2- Blk','IPF-  POW Blk','BPS-BLK','RCB1,RCB2','PBP 1-Mil','200EP POW Blk','375 EP - Mil',
                 '999EP Blk','BP Offset','PC1-135°-Mil','PC2-Mil','PC1- POW Blk','BP- 3x5 Blk','BP-6SC Blk','BP Rubber Gkt','LedLok-5','Concrete Anchor','LedLok-3 5/8','142015','RCB Screws-Bag of 20',
                 'SDS Bag 25','SS W805X100 - Blk','SS AW 30','NC- Blk,CW POW']
    bundles=0
    pallets=0
    boxes=0

    cur_bundle=0
    cur_pallet=0
    cur_box=0


    for key,val in items.items():
        if key in longparts:
            print(key)
            if val > 45:
                new_bundles = round(val/45)
                rem = abs((new_bundles*45) - val)
                bundles += new_bundles
                print(bundles)
                if (cur_bundle+rem) >= 45:
                    bundles += math.ceil((cur_bundle+rem)/45)
                    cur_bundles = 0
                    
                else:
                    cur_bundle += new_bundles
                    
            elif (val+cur_bundle) < 45:
                cur_bundle += val
                
            elif (val+cur_bundle) > 45:
                rem = 45 - cur_bundle
                bundles += math.ceil((val+cur_bundle)/45)
                cur_bundle = val - rem
                
        if key in pallet_parts:
            if key == 'PT1-420- Blk':
                pallets += math.ceil(val/2000)
            elif key == 'PT1-360- Blk':
                pallets += math.ceil(val/2000)
            else:
                if val+cur_pallet < 60:
                    cur_pallet += val
                else:
                    rem = 60 - val
                    pallets += 1
                    cur_pallet= val - rem
        if key in box_parts:
            boxes += 1

    if cur_bundle != 0:
        bundles += 1
    pallets += math.ceil(boxes/30)



    print('approx pallets:',pallets,'\napprox bundles:',bundles)

window = Tk()
window.title('Freight Importer')
window.geometry('300x100')
open_button = Button(master=window,width=20,text='Open File',command=file_name)
import_button = Button(master=window,width=20,text='Transfer Data',command=import_data)
open_button.pack()
import_button.pack()
window.mainloop()
